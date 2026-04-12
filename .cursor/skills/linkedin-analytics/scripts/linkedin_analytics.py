#!/usr/bin/env python3
"""LinkedIn Analytics CLI — parse XLSX exports, store historical data, generate reports."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import unquote

import openpyxl

DATA_DIR = Path(__file__).resolve().parents[4] / "data" / "linkedin-analytics"
POSTS_FILE = DATA_DIR / "posts.jsonl"
ENGAGEMENT_FILE = DATA_DIR / "engagement.jsonl"
FOLLOWERS_FILE = DATA_DIR / "followers.jsonl"
DEMOGRAPHICS_FILE = DATA_DIR / "demographics.json"
SNAPSHOTS_DIR = DATA_DIR / "snapshots"


# ---------------------------------------------------------------------------
# XLSX parsing helpers
# ---------------------------------------------------------------------------

def _parse_date_us(raw: str) -> str:
    """Convert US-format date string (M/D/YYYY) to ISO (YYYY-MM-DD)."""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw.strip()


def _extract_window(ws) -> tuple[str, str]:
    """Extract date window from the DISCOVERY sheet header cell B1."""
    raw = str(ws.cell(1, 2).value or "")
    parts = [p.strip() for p in raw.split("-")]
    if len(parts) == 2:
        return _parse_date_us(parts[0]), _parse_date_us(parts[1])
    return "", ""


def _extract_hashtags(url: str) -> list[str]:
    """Pull hashtag-like tokens from a LinkedIn post URL slug."""
    decoded = unquote(url)
    match = re.search(r"linkedin\.com/posts/[^/]+_(.+?)-(share|ugcPost)", decoded)
    if not match:
        return []
    slug = match.group(1)
    tokens = slug.split("-")
    tags = [t for t in tokens if t and t[0].isupper() and len(t) > 2 and t.isalpha()]
    return [t.lower() for t in tags]


def _safe_int(v) -> int:
    if v is None:
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# XLSX sheet parsers
# ---------------------------------------------------------------------------

def parse_discovery(ws) -> dict:
    window_start, window_end = _extract_window(ws)
    impressions = _safe_int(ws.cell(2, 2).value)
    members_reached = _safe_int(ws.cell(3, 2).value)
    return {
        "window_start": window_start,
        "window_end": window_end,
        "impressions": impressions,
        "members_reached": members_reached,
    }


def parse_engagement(ws) -> list[dict]:
    rows = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if not date_val:
            continue
        rows.append({
            "date": _parse_date_us(str(date_val)),
            "impressions": _safe_int(ws.cell(r, 2).value),
            "engagements": _safe_int(ws.cell(r, 3).value),
        })
    return rows


def parse_top_posts(ws, window_start: str, window_end: str) -> list[dict]:
    """Parse the two side-by-side tables in TOP POSTS and merge them by URL."""
    eng_map: dict[str, dict] = {}
    imp_map: dict[str, dict] = {}

    for r in range(4, ws.max_row + 1):
        # Left table: A=URL, B=date, C=engagements
        url_eng = ws.cell(r, 1).value
        if url_eng:
            eng_map[str(url_eng)] = {
                "post_date": _parse_date_us(str(ws.cell(r, 2).value or "")),
                "engagements": _safe_int(ws.cell(r, 3).value),
            }
        # Right table: E=URL, F=date, G=impressions
        url_imp = ws.cell(r, 5).value
        if url_imp:
            imp_map[str(url_imp)] = {
                "post_date": _parse_date_us(str(ws.cell(r, 6).value or "")),
                "impressions": _safe_int(ws.cell(r, 7).value),
            }

    all_urls = set(eng_map.keys()) | set(imp_map.keys())
    now = datetime.now(UTC).isoformat(timespec="seconds")

    posts = []
    for url in all_urls:
        eng_data = eng_map.get(url, {})
        imp_data = imp_map.get(url, {})
        post_date = eng_data.get("post_date") or imp_data.get("post_date", "")
        impressions = imp_data.get("impressions", 0)
        engagements = eng_data.get("engagements", 0)
        rate = engagements / impressions if impressions > 0 else 0.0

        posts.append({
            "post_url": url,
            "post_date": post_date,
            "impressions": impressions,
            "engagements": engagements,
            "engagement_rate": round(rate, 4),
            "window_start": window_start,
            "window_end": window_end,
            "imported_at": now,
            "hashtags": _extract_hashtags(url),
        })

    posts.sort(key=lambda p: p["impressions"], reverse=True)
    return posts


def parse_followers(ws) -> tuple[int | None, list[dict]]:
    total_raw = ws.cell(1, 2).value
    total = _safe_int(total_raw) if total_raw else None

    rows = []
    for r in range(4, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if not date_val:
            continue
        rows.append({
            "date": _parse_date_us(str(date_val)),
            "new_followers": _safe_int(ws.cell(r, 2).value),
        })

    if total and rows:
        rows[-1]["total_followers"] = total

    return total, rows


def parse_demographics(ws) -> dict:
    categories: dict[str, list[dict]] = {}
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        pct = ws.cell(r, 3).value
        if not cat or not val:
            continue
        key = str(cat).lower().replace(" ", "_")
        categories.setdefault(key, []).append({
            "name": str(val),
            "percentage": str(pct),
        })
    return categories


# ---------------------------------------------------------------------------
# JSONL helpers
# ---------------------------------------------------------------------------

def _read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    records = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if line:
            records.append(json.loads(line))
    return records


def _write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _upsert_daily(existing: list[dict], new_rows: list[dict]) -> list[dict]:
    """Merge daily rows, keeping the latest value for each date."""
    by_date = {r["date"]: r for r in existing}
    for row in new_rows:
        by_date[row["date"]] = row
    return sorted(by_date.values(), key=lambda r: r["date"])


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_import_xlsx(args: argparse.Namespace) -> None:
    filepath = Path(args.file).resolve()
    if not filepath.exists():
        print(f"Error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = {ws.title: ws for ws in wb.worksheets}

    discovery = parse_discovery(sheets["DISCOVERY"])
    window_start = discovery["window_start"]
    window_end = discovery["window_end"]

    engagement_rows = parse_engagement(sheets["ENGAGEMENT"])
    posts = parse_top_posts(sheets["TOP POSTS"], window_start, window_end)
    total_followers, follower_rows = parse_followers(sheets["FOLLOWERS"])
    demographics = parse_demographics(sheets["DEMOGRAPHICS"])
    wb.close()

    # Save raw snapshot
    SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    snapshot = {
        "source_file": filepath.name,
        "imported_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "discovery": discovery,
        "engagement_days": len(engagement_rows),
        "posts_count": len(posts),
        "total_followers": total_followers,
        "follower_days": len(follower_rows),
    }
    snap_name = f"{window_start}_to_{window_end}.json"
    (SNAPSHOTS_DIR / snap_name).write_text(json.dumps(snapshot, indent=2, ensure_ascii=False))

    # Upsert engagement
    existing_eng = _read_jsonl(ENGAGEMENT_FILE)
    merged_eng = _upsert_daily(existing_eng, engagement_rows)
    _write_jsonl(ENGAGEMENT_FILE, merged_eng)

    # Upsert followers
    existing_fol = _read_jsonl(FOLLOWERS_FILE)
    merged_fol = _upsert_daily(existing_fol, follower_rows)
    _write_jsonl(FOLLOWERS_FILE, merged_fol)

    # Append posts (window-tagged, deduplicate by url+window)
    existing_posts = _read_jsonl(POSTS_FILE)
    existing_keys = {
        (p["post_url"], p["window_start"], p["window_end"]) for p in existing_posts
    }
    new_posts = [
        p for p in posts
        if (p["post_url"], p["window_start"], p["window_end"]) not in existing_keys
    ]
    all_posts = existing_posts + new_posts
    _write_jsonl(POSTS_FILE, all_posts)

    # Demographics
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    demographics["exported_at"] = window_end
    DEMOGRAPHICS_FILE.write_text(json.dumps(demographics, indent=2, ensure_ascii=False))

    print(f"Imported: {filepath.name}")
    print(f"  Window: {window_start} to {window_end}")
    print(f"  Discovery: {discovery['impressions']:,} impressions, {discovery['members_reached']:,} reached")
    print(f"  Engagement: {len(engagement_rows)} daily rows ({len(merged_eng)} total stored)")
    print(f"  Posts: {len(posts)} found, {len(new_posts)} new ({len(all_posts)} total stored)")
    print(f"  Followers: {total_followers or 'n/a'} total, {len(follower_rows)} daily rows")
    print(f"  Demographics: {sum(len(v) for v in demographics.values() if isinstance(v, list))} entries")


def cmd_report(args: argparse.Namespace) -> None:
    engagement = _read_jsonl(ENGAGEMENT_FILE)
    posts = _read_jsonl(POSTS_FILE)
    followers = _read_jsonl(FOLLOWERS_FILE)

    if not engagement:
        print("No engagement data. Run import-xlsx first.", file=sys.stderr)
        sys.exit(1)

    days = args.days
    if days:
        cutoff = datetime.now(UTC).strftime("%Y-%m-%d")
        engagement = [e for e in engagement if e["date"] <= cutoff]
        engagement = engagement[-days:]

    total_imp = sum(e["impressions"] for e in engagement)
    total_eng = sum(e["engagements"] for e in engagement)
    eng_rate = total_eng / total_imp if total_imp else 0
    period_days = len(engagement)

    # Widest-window posts for ranking
    widest_posts = _get_widest_window_posts(posts)
    widest_posts.sort(key=lambda p: p["impressions"], reverse=True)

    # Posting frequency from posts with unique dates
    post_dates = sorted({p["post_date"] for p in widest_posts if p["post_date"] >= engagement[0]["date"]})
    posts_per_week = len(post_dates) / (period_days / 7) if period_days > 0 else 0

    # Follower growth
    if followers:
        total_new = sum(f["new_followers"] for f in followers[-period_days:] if "new_followers" in f)
        latest_total = next((f["total_followers"] for f in reversed(followers) if f.get("total_followers")), None)
    else:
        total_new = 0
        latest_total = None

    print("# LinkedIn Performance Report")
    print()
    print(f"**Period**: {engagement[0]['date']} to {engagement[-1]['date']} ({period_days} days)")
    print()
    print("## Summary")
    print()
    print(f"| Metric | Value |")
    print(f"|--------|-------|")
    print(f"| Total impressions | {total_imp:,} |")
    print(f"| Total engagements | {total_eng:,} |")
    print(f"| Engagement rate | {eng_rate:.2%} |")
    print(f"| Avg daily impressions | {total_imp // period_days:,} |")
    if latest_total:
        print(f"| Total followers | {latest_total:,} |")
    print(f"| New followers (period) | {total_new:,} |")
    print(f"| Posts published | {len(post_dates)} |")
    print(f"| Posting frequency | {posts_per_week:.1f} posts/week |")
    print()

    print("## Top Posts by Impressions")
    print()
    print("| # | Date | Impressions | Engagements | Rate | URL |")
    print("|---|------|-------------|-------------|------|-----|")
    for i, p in enumerate(widest_posts[:10], 1):
        rate_str = f"{p['engagement_rate']:.1%}" if p["engagement_rate"] else "0%"
        short_url = p["post_url"].split("?")[0][-60:]
        print(f"| {i} | {p['post_date']} | {p['impressions']:,} | {p['engagements']:,} | {rate_str} | [link]({p['post_url']}) |")
    print()

    # Weekly engagement trend
    print("## Weekly Engagement Trend")
    print()
    weeks: dict[str, dict] = {}
    for e in engagement:
        week = datetime.strptime(e["date"], "%Y-%m-%d").strftime("%Y-W%W")
        w = weeks.setdefault(week, {"impressions": 0, "engagements": 0, "days": 0})
        w["impressions"] += e["impressions"]
        w["engagements"] += e["engagements"]
        w["days"] += 1

    print("| Week | Impressions | Engagements | Avg daily |")
    print("|------|-------------|-------------|-----------|")
    for week, w in sorted(weeks.items()):
        avg = w["impressions"] // w["days"] if w["days"] else 0
        print(f"| {week} | {w['impressions']:,} | {w['engagements']:,} | {avg:,} |")
    print()


def cmd_top_posts(args: argparse.Namespace) -> None:
    posts = _read_jsonl(POSTS_FILE)
    if not posts:
        print("No post data. Run import-xlsx first.", file=sys.stderr)
        sys.exit(1)

    widest = _get_widest_window_posts(posts)

    sort_key = args.sort
    widest.sort(key=lambda p: p.get(sort_key, 0), reverse=True)

    limit = args.limit
    print(f"Top {limit} posts by {sort_key}:")
    print()
    print("| # | Date | Impressions | Engagements | Rate | URL |")
    print("|---|------|-------------|-------------|------|-----|")
    for i, p in enumerate(widest[:limit], 1):
        rate_str = f"{p['engagement_rate']:.1%}" if p.get("engagement_rate") else "0%"
        print(f"| {i} | {p['post_date']} | {p['impressions']:,} | {p['engagements']:,} | {rate_str} | [link]({p['post_url']}) |")


def cmd_trend(args: argparse.Namespace) -> None:
    metric = args.metric
    days = args.days

    if metric == "followers":
        data = _read_jsonl(FOLLOWERS_FILE)
        val_key = "new_followers"
    else:
        data = _read_jsonl(ENGAGEMENT_FILE)
        val_key = metric

    if not data:
        print(f"No {metric} data. Run import-xlsx first.", file=sys.stderr)
        sys.exit(1)

    if days:
        data = data[-days:]

    print(f"## {metric.title()} Trend ({len(data)} days)")
    print()
    print(f"| Date | {metric.title()} |")
    print(f"|------|{'─' * max(len(metric), 6)}|")
    for row in data:
        val = row.get(val_key, 0)
        bar = "█" * min(val // 10, 50) if val > 0 else "·"
        print(f"| {row['date']} | {val:>6,} {bar} |")


def cmd_demographics(args: argparse.Namespace) -> None:
    if not DEMOGRAPHICS_FILE.exists():
        print("No demographics data. Run import-xlsx first.", file=sys.stderr)
        sys.exit(1)

    demo = json.loads(DEMOGRAPHICS_FILE.read_text())
    exported = demo.pop("exported_at", "unknown")
    print(f"# Audience Demographics (as of {exported})")
    print()

    display_order = ["seniority", "industry", "location", "company_size", "job_title", "company"]
    for cat in display_order:
        entries = demo.get(cat, [])
        if not entries:
            continue
        title = cat.replace("_", " ").title()
        print(f"## {title}")
        print()
        print(f"| {title} | % |")
        print(f"|{'─' * max(len(title), 20)}|-----|")
        for entry in entries:
            print(f"| {entry['name']} | {entry['percentage']} |")
        print()


def cmd_compare(args: argparse.Namespace) -> None:
    file1 = Path(args.file1).resolve()
    file2 = Path(args.file2).resolve()

    for f in (file1, file2):
        if not f.exists():
            print(f"Error: file not found: {f}", file=sys.stderr)
            sys.exit(1)

    wb1 = openpyxl.load_workbook(file1, read_only=True, data_only=True)
    wb2 = openpyxl.load_workbook(file2, read_only=True, data_only=True)

    d1 = parse_discovery(wb1["DISCOVERY"])
    d2 = parse_discovery(wb2["DISCOVERY"])

    _, fol1 = parse_followers(wb1["FOLLOWERS"])
    total1_val = _safe_int(wb1["FOLLOWERS"].cell(1, 2).value)
    _, fol2 = parse_followers(wb2["FOLLOWERS"])
    total2_val = _safe_int(wb2["FOLLOWERS"].cell(1, 2).value)

    p1 = parse_top_posts(wb1["TOP POSTS"], d1["window_start"], d1["window_end"])
    p2 = parse_top_posts(wb2["TOP POSTS"], d2["window_start"], d2["window_end"])

    wb1.close()
    wb2.close()

    print("# Export Comparison")
    print()
    print(f"| Metric | {file1.name[:40]} | {file2.name[:40]} | Delta |")
    print("|--------|" + "-" * 42 + "|" + "-" * 42 + "|-------|")

    imp_delta = d2["impressions"] - d1["impressions"]
    print(f"| Window | {d1['window_start']} – {d1['window_end']} | {d2['window_start']} – {d2['window_end']} | |")
    print(f"| Impressions | {d1['impressions']:,} | {d2['impressions']:,} | {imp_delta:+,} |")
    print(f"| Members reached | {d1['members_reached']:,} | {d2['members_reached']:,} | {d2['members_reached'] - d1['members_reached']:+,} |")

    if total1_val and total2_val:
        print(f"| Total followers | {total1_val:,} | {total2_val:,} | {total2_val - total1_val:+,} |")

    new_fol1 = sum(f["new_followers"] for f in fol1)
    new_fol2 = sum(f["new_followers"] for f in fol2)
    print(f"| New followers | {new_fol1} | {new_fol2} | {new_fol2 - new_fol1:+} |")
    print()

    # Posts that appear in both: compare impressions
    urls1 = {p["post_url"]: p for p in p1}
    urls2 = {p["post_url"]: p for p in p2}
    common = set(urls1.keys()) & set(urls2.keys())

    if common:
        print("## Posts in Both Exports (impression changes)")
        print()
        print("| Date | Impressions (1) | Impressions (2) | Delta | URL |")
        print("|------|-----------------|-----------------|-------|-----|")
        rows = []
        for url in common:
            delta = urls2[url]["impressions"] - urls1[url]["impressions"]
            rows.append((urls1[url]["post_date"], urls1[url]["impressions"], urls2[url]["impressions"], delta, url))
        rows.sort(key=lambda r: abs(r[3]), reverse=True)
        for date, imp1, imp2, delta, url in rows:
            print(f"| {date} | {imp1:,} | {imp2:,} | {delta:+,} | [link]({url}) |")
        print()

    only1 = set(urls1.keys()) - set(urls2.keys())
    only2 = set(urls2.keys()) - set(urls1.keys())
    if only2:
        print(f"## Posts only in export 2 ({len(only2)})")
        print()
        for url in sorted(only2):
            p = urls2[url]
            print(f"- {p['post_date']}: {p['impressions']:,} imp — [link]({url})")
        print()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_widest_window_posts(posts: list[dict]) -> list[dict]:
    """For each post URL, keep the record from the widest measurement window."""
    best: dict[str, dict] = {}
    for p in posts:
        url = p["post_url"]
        if url not in best:
            best[url] = p
            continue
        existing = best[url]
        existing_span = _window_days(existing)
        new_span = _window_days(p)
        if new_span > existing_span:
            best[url] = p
    return list(best.values())


def _window_days(p: dict) -> int:
    try:
        s = datetime.strptime(p["window_start"], "%Y-%m-%d")
        e = datetime.strptime(p["window_end"], "%Y-%m-%d")
        return (e - s).days
    except (ValueError, KeyError):
        return 0


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="linkedin_analytics",
        description="LinkedIn analytics CLI — parse XLSX exports and generate reports",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # import-xlsx
    p_import = sub.add_parser("import-xlsx", help="Import a LinkedIn XLSX analytics export")
    p_import.add_argument("file", help="Path to the XLSX file")

    # report
    p_report = sub.add_parser("report", help="Generate a performance report")
    p_report.add_argument("--days", type=int, default=None, help="Limit to last N days")

    # top-posts
    p_top = sub.add_parser("top-posts", help="List top posts by metric")
    p_top.add_argument("--limit", type=int, default=10, help="Number of posts to show")
    p_top.add_argument("--sort", choices=["impressions", "engagements", "engagement_rate"], default="impressions")

    # trend
    p_trend = sub.add_parser("trend", help="Show daily metric trends")
    p_trend.add_argument("--metric", choices=["impressions", "engagements", "followers"], default="impressions")
    p_trend.add_argument("--days", type=int, default=None, help="Limit to last N days")

    # demographics
    sub.add_parser("demographics", help="Show audience demographics")

    # compare
    p_cmp = sub.add_parser("compare", help="Compare two XLSX exports")
    p_cmp.add_argument("file1", help="First XLSX file")
    p_cmp.add_argument("file2", help="Second XLSX file")

    args = parser.parse_args()

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)

    commands = {
        "import-xlsx": cmd_import_xlsx,
        "report": cmd_report,
        "top-posts": cmd_top_posts,
        "trend": cmd_trend,
        "demographics": cmd_demographics,
        "compare": cmd_compare,
    }
    commands[args.command](args)


if __name__ == "__main__":
    main()
